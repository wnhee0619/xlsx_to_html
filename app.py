from flask import Flask, render_template, request, send_file, redirect, url_for
import openpyxl
import io

app = Flask(__name__)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/generate', methods=['POST'])
def generate_table():
    # 파일 업로드 처리
    file = request.files['file']
    file.save('uploaded_file.xlsx')
    
    # 엑셀 파일 로드
    workbook = openpyxl.load_workbook('uploaded_file.xlsx')
    
    # 시트 이름 가져오기
    sheet_name = request.form['sheet_name']
    worksheet = workbook[sheet_name]
    
    # 병합된 셀 범위 가져오기
    merged_cells_ranges = list(worksheet.merged_cells.ranges)
    
    html = "<table>\n"
    
    # 헤더 생성
    header_row = next(worksheet.iter_rows())
    html += "  <tr>\n"
    for cell in header_row:
        colSpan = 1
        rowSpan = 1
        
        # 병합된 셀인지 확인하고, 병합된 셀의 크기 파악
        for merged_cell_range in merged_cells_ranges:
            if cell.coordinate in merged_cell_range:
                colSpan = merged_cell_range.max_col - merged_cell_range.min_col + 1
                rowSpan = merged_cell_range.max_row - merged_cell_range.min_row + 1
                if cell.coordinate == merged_cell_range.coord or cell.value is not None:
                    html += f"    <th colspan={colSpan} rowspan={rowSpan}>{cell.value if cell.value else ''}</th>\n"             
                break
        
        if cell.coordinate not in merged_cell_range:
            html += f"    <th colspan={colSpan} rowspan={rowSpan}>{cell.value if cell.value else ''}</th>\n"
    html += "  </tr>\n"
    
    # 데이터 생성
    for row in worksheet.iter_rows(min_row=2):
        html += "  <tr>\n"
        for cell in row:
            colSpan = 1
            rowSpan = 1
            
            # 병합된 셀인지 확인하고, 병합된 셀의 크기 파악
            for merged_cell_range in merged_cells_ranges:
                if cell.coordinate in merged_cell_range:
                    colSpan = merged_cell_range.max_col - merged_cell_range.min_col + 1
                    rowSpan = merged_cell_range.max_row - merged_cell_range.min_row + 1
                    break
                
            if cell.coordinate == merged_cells_ranges[0].coord or cell.value is not None:
                html += f"    <td colspan={colSpan} rowspan={rowSpan}>{cell.value if cell.value else ''}</td>\n"
            elif cell.coordinate not in merged_cell_range:
                html += "    <td></td>\n"
        html += "  </tr>\n"
    
    html += "</table>"
    
    # HTML을 텍스트 파일로 변환하여 다운로드
    output = io.BytesIO()
    output.write(html.encode('utf-8'))
    output.seek(0)
    
    return redirect(url_for('download_result', result=html))

@app.route('/download')
def download_result():
    html = request.args.get('result')
    output = io.BytesIO()
    output.write(html.encode('utf-8'))
    output.seek(0)
    return send_file(output, as_attachment=True, download_name='result.txt', mimetype='text/plain')

if __name__ == '__main__':
    app.run(debug=True)
